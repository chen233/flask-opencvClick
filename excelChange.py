import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font  # 改为使用字体颜色
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import chardet
from datetime import datetime
import os


def get_file_encoding(file_path):
    """自动检测文件编码（解决中文解码问题）"""
    with open(file_path, 'rb') as f:
        raw_data = f.read(1024)
        result = chardet.detect(raw_data)
        encoding = result['encoding'] or 'gbk'  # 检测失败默认用GBK
        if encoding.lower() in ['gb2312', 'gbk']:
            encoding = 'gbk'
        return encoding


def extract_date_from_txt(txt_path):
    """从TXT文件名提取日期（格式：2025-09-21_07-02-26.txt → 2025/09/21）"""
    txt_filename = os.path.basename(txt_path)
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', txt_filename)
    if date_match:
        return date_match.group(1).replace('-', '/')  # 转为YYYY/MM/DD
    return datetime.now().strftime('%Y/%m/%d')  # 默认当前日期


def parse_date(date_str):
    """兼容多种日期格式（YYYY/MM/DD、YYYY-MM-DD、YYYY-MM-DD HH:MM:SS等）"""
    # 尝试多种常见格式解析
    date_formats = [
        '%Y/%m/%d',  # 2025/09/21
        '%Y-%m-%d',  # 2025-09-21
        '%Y-%m-%d %H:%M:%S',  # 2025-09-21 00:00:00
        '%Y/%m/%d %H:%M:%S'  # 2025/09/21 00:00:00
    ]
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt)
        except:
            continue
    # 所有格式都失败时，返回当前日期（避免计算错误）
    print(f"警告：无法解析日期格式 '{date_str}'，默认使用当前日期")
    return datetime.now()


def calculate_days(start_date_str, end_date_str):
    """计算两个日期的天数差（支持多格式日期）"""
    # 先统一解析为datetime对象
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    # 计算天数差（取绝对值）
    return abs((end_date - start_date).days)


def txt_to_excel_with_history(txt_path, excel_path):
    """支持手动修改日期后重新计算天数，使用文字颜色标记可改动区域"""
    excel_columns = [
        '角色ID', '最初统计时间', '更新统计时间', '初始金币', '更新金币',
        '天数', '每天产量', '总计产生', '银币', '等级', '门派',
        '角色', '服务器', '角色名', '关联账号', '关联密码'
    ]

    # ---------------------- 1. 读取TXT数据 ----------------------
    file_encoding = get_file_encoding(txt_path)
    try:
        with open(txt_path, 'r', encoding=file_encoding, errors='ignore') as f:
            lines = f.readlines()
    except:
        with open(txt_path, 'r', encoding='gbk', errors='ignore') as f:
            lines = f.readlines()

    header = re.split(r'-+', lines[0].strip())
    txt_core_data = []
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        row = re.split(r'-+', line)
        if len(row) < len(header):
            row += [''] * (len(header) - len(row))
        item = dict(zip(header, row))
        txt_core_data.append({
            '角色ID': item.get('角色ID', '').strip(),
            '更新统计时间': extract_date_from_txt(txt_path),
            '更新金币': item.get('金币', '0').strip(),
            '银币': item.get('银币', '0').strip(),
            '等级': item.get('等级', '').strip(),
            '门派': item.get('门派', '').strip(),
            '角色': item.get('角色', '').strip(),
            '服务器': item.get('服务器', '').strip(),
            '角色名': item.get('角色名', '').strip(),
            '关联账号': item.get('关联账号', '').strip(),
            '关联密码': item.get('关联密码', '').strip()
        })

    # ---------------------- 2. 读取已有Excel数据 ----------------------
    if os.path.exists(excel_path):
        try:
            excel_df = pd.read_excel(excel_path, sheet_name='角色数据')
            # 补全缺失列
            for col in excel_columns:
                if col not in excel_df.columns:
                    excel_df[col] = ''
            excel_df = excel_df[excel_columns]
        except Exception as e:
            print(f"读取Excel失败（{e}），将新建Excel文件")
            excel_df = pd.DataFrame(columns=excel_columns)
    else:
        excel_df = pd.DataFrame(columns=excel_columns)

    # ---------------------- 3. 增量更新数据（强制重新计算天数） ----------------------
    id_to_index = {str(row['角色ID']).strip(): idx for idx, row in excel_df.iterrows()}

    for new_data in txt_core_data:
        role_id = new_data['角色ID']
        if role_id not in id_to_index:
            # 首次录入
            print(f"首次录入角色ID：{role_id}")
            new_row = {
                '角色ID': role_id,
                '最初统计时间': new_data['更新统计时间'],
                '更新统计时间': new_data['更新统计时间'],
                '初始金币': new_data['更新金币'],
                '更新金币': new_data['更新金币'],
                '银币': new_data['银币'],
                '等级': new_data['等级'],
                '门派': new_data['门派'],
                '角色': new_data['角色'],
                '服务器': new_data['服务器'],
                '角色名': new_data['角色名'],
                '关联账号': new_data['关联账号'],
                '关联密码': new_data['关联密码']
            }
            # 计算天数（首次录入可能为0，默认1）
            days = calculate_days(new_row['最初统计时间'], new_row['更新统计时间'])
            new_row['天数'] = days if days > 0 else 1
            total_produce = float(new_row['更新金币']) - float(new_row['初始金币'])
            new_row['总计产生'] = round(total_produce, 2)
            new_row['每天产量'] = round(total_produce / new_row['天数'], 2) if new_row['天数'] != 0 else 0
            # 添加到DataFrame
            excel_df = pd.concat([excel_df, pd.DataFrame([new_row])], ignore_index=True)
            id_to_index[role_id] = len(excel_df) - 1
        else:
            # 已存在角色：更新数据并强制重新计算天数
            print(f"更新角色ID：{role_id}（强制重新计算天数）")
            idx = id_to_index[role_id]
            # 更新可变动字段
            excel_df.at[idx, '更新统计时间'] = new_data['更新统计时间']
            excel_df.at[idx, '更新金币'] = new_data['更新金币']
            excel_df.at[idx, '银币'] = new_data['银币']
            excel_df.at[idx, '等级'] = new_data['等级']
            excel_df.at[idx, '门派'] = new_data['门派']
            excel_df.at[idx, '角色'] = new_data['角色']
            excel_df.at[idx, '服务器'] = new_data['服务器']
            excel_df.at[idx, '角色名'] = new_data['角色名']
            excel_df.at[idx, '关联账号'] = new_data['关联账号']
            excel_df.at[idx, '关联密码'] = new_data['关联密码']

            # 强制重新计算（即使手动修改过最初统计时间也会生效）
            start_date = excel_df.at[idx, '最初统计时间']  # 可能是手动修改后的日期
            end_date = excel_df.at[idx, '更新统计时间']
            days = calculate_days(start_date, end_date)  # 关键：用新的日期解析函数
            excel_df.at[idx, '天数'] = days if days > 0 else 1
            total_produce = float(excel_df.at[idx, '更新金币']) - float(excel_df.at[idx, '初始金币'])
            excel_df.at[idx, '总计产生'] = round(total_produce, 2)
            excel_df.at[idx, '每天产量'] = round(total_produce / excel_df.at[idx, '天数'], 2) if excel_df.at[
                                                                                                     idx, '天数'] != 0 else 0

    # 处理空值
    final_df = excel_df.fillna({
        '初始金币': 0, '更新金币': 0, '天数': 0, '每天产量': 0, '总计产生': 0, '银币': 0
    })

    # ---------------------- 4. 写入Excel并设置文字颜色标记 ----------------------
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if '角色数据' in wb.sheetnames:
            wb.remove(wb['角色数据'])
        ws = wb.create_sheet(title='角色数据')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '角色数据'

    # 定义文字颜色（红色：可改动字段；绿色：自动计算字段）
    red_font = Font(color="FF0000")  # 红色文字
    green_font = Font(color="008000")  # 绿色文字（使用较深的绿色以确保可读性）
    default_font = Font(color="000000")  # 默认黑色文字

    # 写入数据（含表头）
    for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            # 设置单元格值
            if isinstance(value, (int, float)):
                ws.cell(row=r_idx, column=c_idx, value=round(value, 2))
            else:
                ws.cell(row=r_idx, column=c_idx, value=str(value).strip())

            # 保持表头为默认黑色
            if r_idx == 1:
                ws.cell(row=r_idx, column=c_idx).font = default_font

    # 建立“列名→列号”映射（表头在第1行）
    col_map = {}
    for c_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=c_idx).value
        if col_name in excel_columns:
            col_map[col_name] = c_idx

    # 数据行范围（从第2行到最后一行）
    data_start_row = 2
    data_end_row = ws.max_row

    # 红色文字：更新金币（可手动修改的字段）
    if '更新金币' in col_map:
        red_col_idx = col_map['更新金币']
        for row in range(data_start_row, data_end_row + 1):
            ws.cell(row=row, column=red_col_idx).font = red_font

    # 绿色文字：自动计算字段（天数、每天产量、总计产生）
    green_cols = ['天数', '每天产量', '总计产生']
    for col_name in green_cols:
        if col_name in col_map:
            green_col_idx = col_map[col_name]
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=green_col_idx).font = green_font

    # 保存文件
    wb.save("角色数据_历史追踪.xlsx")
    print(f"\n✅ 数据更新完成！")
    print("提示：")
    print("- 🔴 红色文字（更新金币）：可手动修改的字段")
    print("- 🟢 绿色文字（天数、每天产量、总计产生）：自动计算的字段")
    print("- 手动修改日期后，重新运行本程序即可更新天数！")


